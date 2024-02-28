from openpyxl import Workbook,load_workbook


# ws.max_row,ws.max_column
# ws.values
# a = ws.values
#next(a)
# b= ws.columns
# next(b)[0].value
def load_excel(name):
    return load_workbook(name)
def extract_sheet_names(wb):
    return wb.sheetnames
def selected_sheet(wb):
    pass
def extract_column_names(ws):
    col = []
    (ws.max_column)
    count =0
    gen = ws.columns
    while count < ws.max_column:
        col.append(next(gen)[0].value)
        count+=1
    print(col)
    return col
    
def delete_column_excel(ws,index):
    print('delete_column_excel-----',index)
    ws.delete_cols(index)
    return True
def change_sheet_active():
    pass
def sort_sheet():
    pass
def filter_sheet():
    pass

def font_size_change():
    pass
def space_modifier():
    pass

