from django.shortcuts import render,redirect,HttpResponseRedirect
from django.urls import reverse
from .form import UploadFileForm
from .models import ExcelFile
from django.shortcuts import get_object_or_404
from openpyxl import Workbook, load_workbook
from django.http.response import JsonResponse,HttpResponseBadRequest
from .utils import extract_column_names,delete_column_excel
def upload_and_extract(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST,request.FILES)
        print(form.data,form.files)
        if form.is_valid() :
            print('valid')
            saved = form.save()
            return HttpResponseRedirect(reverse('uploaded_extracting',args=[saved.id]))
        else:
            print('invalid')
        return render(request,'excel_reduction/uploaded.html',{'form':form})
    if request.method == 'GET':
        form = UploadFileForm()
        return render(request,'excel_reduction/home.html',{'form':form})

def uploaded_extracting(request,id,*args, **kwargs):
    print(id)
    print(request.GET)
    instance = ExcelFile.objects.get(id=id)
    wb = load_workbook(instance.file_data.path)
    print(instance.file_data.file)

    # now sheet ,column option , peaking option , delete option ,confirmation, scheduling,reminder
    if request.GET.get('sheet'):
        pass
    # ''' here all data is extracted and then passed to frontend from which user can decide what to do'''
    ws = wb.active
    request.session['excel_selected'] = id
    print(ws)
    context = {'sheetnames':wb.sheetnames,'sheet_columns':{}}
    
    return render(request,'excel_reduction/uploaded_extracting.html',context)

# Workbook.get_sheet_by_name()

def extraction_api(request):
    '''Just send column for now to frontend '''
    print('session',request.session['excel_selected']) # get file from session
    if not request.session['excel_selected']:
        return HttpResponseBadRequest()
    instance = ExcelFile.objects.get(id=int(request.session['excel_selected']))
    wb = load_workbook(instance.file_data.path)
    ws = wb.active
    col = extract_column_names(ws)
    if request.GET.get('sheet'):
        sheet = request.GET.get('sheet')
        if request.GET.get('column'):
            column = request.GET.get('column')
            print('got column')
            return JsonResponse({'columns':'received'})
        return JsonResponse({'sheet_selected':sheet,'columns':col})

def delete_column(request):
    """JSONresponse method , query using jqery only ,show all data of excel and on selection delete specific columns One idea is to create a profile for specific file so that we can directly select what is to be done with a kind of file """
    if request.GET.get('sheet'):
            sheet = request.GET.get('sheet')
            print(request.GET.getlist('column'))
            if request.GET.getlist('column'):
                instance = ExcelFile.objects.get(id=int(request.session['excel_selected']))
                wb = load_workbook(instance.file_data.path)
                ws = wb.get_sheet_by_name(sheet)
                print(extract_column_names(ws))
                column = sorted(list(map(int,request.GET.getlist('column'))),reverse=True)
                delete = [ delete_column_excel(ws,i) for i in column]
                print(extract_column_names(ws))
                wb.save(instance.file_data.name)
                
                return JsonResponse({'columns':'received','status':'deleted'})
    return JsonResponse({'status':'error',})